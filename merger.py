from flask import Flask, request, jsonify, make_response
import logging
import base64
from io import BytesIO
from openpyxl import load_workbook, Workbook

app = Flask(__name__)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@app.route('/merge-excel', methods=['POST'])
def merge_excel():
    logger.info('Processing Excel merge request')

    try:
        # Get JSON payload
        req_body = request.get_json()
        if not req_body:
            return make_response(jsonify({"error": "❌ Invalid JSON payload"}), 400)

        files = req_body.get("files", [])
        if not files:
            return make_response(jsonify({"error": "❌ No files provided in payload"}), 400)

        merged_wb = Workbook()
        merged_wb.remove(merged_wb.active)  # Remove default sheet

        for file in files:
            file_name = file.get("name")
            content_b64 = file.get("content")
            if not file_name or not content_b64:
                logger.warning(f"Skipping file due to missing name or content: {file_name}")
                continue

            # Validate file extension
            if not file_name.lower().endswith('.xlsx'):
                logger.warning(f"Skipping non-Excel file: {file_name}")
                continue

            try:
                file_bytes = base64.b64decode(content_b64)
                wb = load_workbook(BytesIO(file_bytes))
            except Exception as e:
                logger.error(f"Failed to process file {file_name}: {str(e)}")
                continue

            for sheet_name in wb.sheetnames:
                source_sheet = wb[sheet_name]
                new_sheet_title = sheet_name
                # Avoid duplicate sheet names
                copy_suffix = 1
                while new_sheet_title in merged_wb.sheetnames:
                    new_sheet_title = f"{sheet_name}_copy{copy_suffix}"
                    copy_suffix += 1
                new_sheet = merged_wb.create_sheet(title=new_sheet_title)
                for row in source_sheet.iter_rows(values_only=True):
                    new_sheet.append(row)

        if not merged_wb.sheetnames:
            return make_response(jsonify({"error": "❌ No valid sheets merged"}), 400)

        # Save merged workbook to BytesIO
        output_stream = BytesIO()
        merged_wb.save(output_stream)
        output_stream.seek(0)
        result_b64 = base64.b64encode(output_stream.read()).decode("utf-8")

        return jsonify({"merged_file": result_b64})

    except Exception as e:
        logger.error(f"❌ Internal error: {str(e)}")
        return make_response(jsonify({"error": f"Internal Server Error: {str(e)}"}), 500)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)