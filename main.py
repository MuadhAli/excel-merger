from openpyxl import load_workbook, Workbook
import os

def merge_excels_with_all_sheets(folder_path, output_file):
    merged_wb = Workbook()
    merged_wb.remove(merged_wb.active)

    # Get all Excel files in the folder
    file_paths = [os.path.join(folder_path, f) for f in os.listdir(folder_path)
                  if f.endswith(".xlsx") and not f.startswith("~$")]

    if not file_paths:
        print("❌ No Excel files found in the folder.")
        return

    for file_path in file_paths:
        print(f"🔄 Processing: {file_path}")
        wb = load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            print(f"   📄 Copying sheet: {sheet_name}")
            source_sheet = wb[sheet_name]

            new_sheet_title = sheet_name
            while new_sheet_title in merged_wb.sheetnames:
                new_sheet_title += "_copy"

            new_sheet = merged_wb.create_sheet(title=new_sheet_title)

            for row in source_sheet.iter_rows(values_only=True):
                new_sheet.append(row)

    merged_wb.save(output_file)
    print(f"✅ Merged file created: {output_file}")

# Example usage
folder_path = "D:\Excel Merger\excel-files"  # Put all your Excel files in this folder
output_file = "merged_output.xlsx"
merge_excels_with_all_sheets(folder_path, output_file)
